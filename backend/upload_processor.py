import csv
import hashlib
import io
import json
import re
from datetime import datetime

from db import get_connection

PROTECTED_TRIAL_COLS = {
    "id", "registry_sources", "all_registry_ids", "ingested_at",
    "raw_snapshot_path", "has_news",
}

ALLOWED_ORG_COLS = {
    "org_type", "therapeutic_focus", "white_label_signal", "funding_stage",
    "offerings", "website", "linkedin_url", "notes", "regions_served",
    "existing_integrations",
}

REQUIRED = {
    "trials": [],          # at least one of id / title_brief enforced in code
    "organizations": ["canonical_name"],
    "contacts": ["full_name", "org_name"],
}


def _parse_rows(file_bytes: bytes, filename: str) -> list[dict]:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext == "xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []
        headers = [_normalize_header(str(h) if h is not None else "") for h in rows[0]]
        return [dict(zip(headers, [str(v) if v is not None else "" for v in row])) for row in rows[1:]]
    else:
        text = file_bytes.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        return [
            {_normalize_header(k): (v or "") for k, v in row.items()}
            for row in reader
        ]


def _normalize_header(h: str) -> str:
    return re.sub(r"\s+", "_", h.strip().lower())


def _jaccard(a: str, b: str) -> float:
    ta = set(re.sub(r"[^\w]", " ", a.lower()).split())
    tb = set(re.sub(r"[^\w]", " ", b.lower()).split())
    if not ta or not tb:
        return 0.0
    return len(ta & tb) / len(ta | tb)


def _upload_id(title: str, sponsor: str) -> str:
    h = hashlib.md5(f"{title}|{sponsor}".encode()).hexdigest()[:8]
    return f"UPLOAD-{h}"


def _get_trial_columns(conn) -> set:
    rows = conn.execute("PRAGMA table_info(trials)").fetchall()
    return {r["name"] for r in rows}


def _process_trials(rows: list, conn) -> dict:
    matched = inserted = skipped = 0
    errors = []
    merge_cands = 0
    preview = []
    all_cols = _get_trial_columns(conn)
    now = datetime.utcnow().isoformat()

    for i, row in enumerate(rows, 1):
        try:
            trial_id = row.get("id", "").strip()
            title = row.get("title_brief", "").strip()
            sponsor = row.get("sponsor", "").strip()

            if not trial_id and not title:
                errors.append({"row": i, "field": "id/title_brief", "message": "Row must have id or title_brief"})
                skipped += 1
                continue

            upload_cols = {
                k: v for k, v in row.items()
                if k in all_cols and k not in PROTECTED_TRIAL_COLS and v.strip()
            }

            if trial_id:
                exists = conn.execute("SELECT id FROM trials WHERE id = ?", (trial_id,)).fetchone()
                if exists:
                    if upload_cols:
                        set_sql = ", ".join(f"{k} = ?" for k in upload_cols)
                        conn.execute(f"UPDATE trials SET {set_sql} WHERE id = ?",
                                     list(upload_cols.values()) + [trial_id])
                    matched += 1
                    if len(preview) < 5:
                        preview.append({"action": "updated", "id": trial_id})
                else:
                    upload_cols["id"] = trial_id
                    upload_cols.setdefault("ingested_at", now)
                    upload_cols.setdefault("registry_sources", '["UPLOAD"]')
                    _insert_trial(conn, upload_cols)
                    inserted += 1
                    if len(preview) < 5:
                        preview.append({"action": "inserted", "id": trial_id})
            else:
                # Fuzzy match
                candidates_q = conn.execute(
                    "SELECT id, title_brief, sponsor FROM trials WHERE LOWER(sponsor) LIKE ? LIMIT 100",
                    (f"%{sponsor.lower()[:30]}%",)
                ).fetchall() if sponsor else []

                best_match = None
                best_score = 0.0
                for c in candidates_q:
                    score = 0.0
                    if title and c["title_brief"]:
                        score += 0.6 * _jaccard(title, c["title_brief"])
                    if sponsor and c["sponsor"]:
                        score += 0.4 * _jaccard(sponsor, c["sponsor"])
                    if score > best_score:
                        best_score = score
                        best_match = c

                if best_match and best_score >= 0.8:
                    if upload_cols:
                        set_sql = ", ".join(f"{k} = ?" for k in upload_cols)
                        conn.execute(f"UPDATE trials SET {set_sql} WHERE id = ?",
                                     list(upload_cols.values()) + [best_match["id"]])
                    matched += 1
                    if len(preview) < 5:
                        preview.append({"action": "fuzzy_updated", "id": best_match["id"], "score": best_score})

                elif best_match and 0.6 <= best_score < 0.8:
                    new_id = _upload_id(title, sponsor)
                    upload_cols["id"] = new_id
                    upload_cols["title_brief"] = title
                    upload_cols["sponsor"] = sponsor
                    upload_cols.setdefault("ingested_at", now)
                    upload_cols.setdefault("registry_sources", '["UPLOAD"]')
                    _insert_trial(conn, upload_cols)
                    _insert_merge_candidate(conn, best_match["id"], new_id, best_score,
                                            ["title_brief", "sponsor"], {"title_brief": best_score * 0.6, "sponsor": best_score * 0.4})
                    inserted += 1
                    merge_cands += 1
                    if len(preview) < 5:
                        preview.append({"action": "inserted_with_candidate", "id": new_id})

                else:
                    new_id = _upload_id(title, sponsor)
                    upload_cols["id"] = new_id
                    upload_cols["title_brief"] = title
                    upload_cols["sponsor"] = sponsor
                    upload_cols.setdefault("ingested_at", now)
                    upload_cols.setdefault("registry_sources", '["UPLOAD"]')
                    _insert_trial(conn, upload_cols)
                    inserted += 1
                    if len(preview) < 5:
                        preview.append({"action": "inserted", "id": new_id})

        except Exception as e:
            errors.append({"row": i, "field": "", "message": str(e)})
            skipped += 1

    conn.commit()
    return {
        "matched": matched, "inserted": inserted, "skipped": skipped,
        "errors": errors, "merge_candidates": merge_cands, "preview": preview,
    }


def _insert_trial(conn, cols: dict):
    k = ", ".join(cols.keys())
    p = ", ".join("?" * len(cols))
    conn.execute(f"INSERT OR REPLACE INTO trials ({k}) VALUES ({p})", list(cols.values()))


def _insert_merge_candidate(conn, a_id, b_id, score, match_fields, match_scores):
    exists = conn.execute(
        """SELECT id FROM merge_candidates
           WHERE entity_type = 'trials'
           AND ((record_a_id = ? AND record_b_id = ?) OR (record_a_id = ? AND record_b_id = ?))""",
        (a_id, b_id, b_id, a_id)
    ).fetchone()
    if not exists:
        conn.execute(
            """INSERT INTO merge_candidates
               (entity_type, record_a_id, record_b_id, confidence, match_fields, match_scores, status, created_at)
               VALUES ('trials', ?, ?, ?, ?, ?, 'PENDING', ?)""",
            (a_id, b_id, score, json.dumps(match_fields), json.dumps(match_scores),
             datetime.utcnow().isoformat()),
        )


def _process_organizations(rows: list, conn) -> dict:
    matched = inserted = skipped = 0
    errors = []
    preview = []
    now = datetime.utcnow().isoformat()

    for i, row in enumerate(rows, 1):
        try:
            canonical = row.get("canonical_name", "").strip()
            if not canonical:
                errors.append({"row": i, "field": "canonical_name", "message": "canonical_name is required"})
                skipped += 1
                continue

            upload_cols = {k: v for k, v in row.items() if k in ALLOWED_ORG_COLS and v.strip()}

            alias = conn.execute(
                "SELECT org_id FROM organization_aliases WHERE alias = ?",
                (canonical.lower(),)
            ).fetchone()

            if alias:
                org_id = alias["org_id"]
                if upload_cols:
                    set_sql = ", ".join(f"{k} = ?" for k in upload_cols)
                    conn.execute(f"UPDATE organizations SET {set_sql} WHERE id = ?",
                                 list(upload_cols.values()) + [org_id])
                matched += 1
                if len(preview) < 5:
                    preview.append({"action": "updated", "id": org_id})
            else:
                from org_extractor import slugify
                slug = slugify(canonical)
                if not slug:
                    skipped += 1
                    continue
                upload_cols["id"] = slug
                upload_cols["canonical_name"] = canonical
                upload_cols.setdefault("created_at", now)
                k = ", ".join(upload_cols.keys())
                p = ", ".join("?" * len(upload_cols))
                conn.execute(f"INSERT OR IGNORE INTO organizations ({k}) VALUES ({p})", list(upload_cols.values()))
                conn.execute(
                    "INSERT OR IGNORE INTO organization_aliases (alias, org_id) VALUES (?, ?)",
                    (canonical.lower(), slug),
                )
                inserted += 1
                if len(preview) < 5:
                    preview.append({"action": "inserted", "id": slug})

        except Exception as e:
            errors.append({"row": i, "field": "", "message": str(e)})
            skipped += 1

    conn.commit()
    return {
        "matched": matched, "inserted": inserted, "skipped": skipped,
        "errors": errors, "merge_candidates": 0, "preview": preview,
    }


def _process_contacts(rows: list, conn) -> dict:
    matched = inserted = skipped = 0
    errors = []
    preview = []
    CONTACT_COLS = {"title", "department", "email", "linkedin_url", "source_url", "is_decision_maker", "notes"}
    now = datetime.utcnow().isoformat()

    for i, row in enumerate(rows, 1):
        try:
            full_name = row.get("full_name", "").strip()
            org_name = row.get("org_name", "").strip()
            if not full_name:
                errors.append({"row": i, "field": "full_name", "message": "full_name is required"})
                skipped += 1
                continue
            if not org_name:
                errors.append({"row": i, "field": "org_name", "message": "org_name is required"})
                skipped += 1
                continue

            alias = conn.execute(
                "SELECT org_id FROM organization_aliases WHERE alias = ?",
                (org_name.lower(),)
            ).fetchone()
            if not alias:
                errors.append({"row": i, "field": "org_name", "message": f'Organization "{org_name}" not found'})
                skipped += 1
                continue

            org_id = alias["org_id"]
            upload_cols = {k: v for k, v in row.items() if k in CONTACT_COLS and v.strip()}

            existing = conn.execute(
                "SELECT id FROM org_contacts WHERE LOWER(full_name) = ? AND org_id = ?",
                (full_name.lower(), org_id),
            ).fetchone()

            if existing:
                if upload_cols:
                    set_sql = ", ".join(f"{k} = ?" for k in upload_cols)
                    conn.execute(f"UPDATE org_contacts SET {set_sql} WHERE id = ?",
                                 list(upload_cols.values()) + [existing["id"]])
                matched += 1
                if len(preview) < 5:
                    preview.append({"action": "updated", "name": full_name})
            else:
                upload_cols["full_name"] = full_name
                upload_cols["org_id"] = org_id
                upload_cols.setdefault("created_at", now)
                k = ", ".join(upload_cols.keys())
                p = ", ".join("?" * len(upload_cols))
                conn.execute(f"INSERT INTO org_contacts ({k}) VALUES ({p})", list(upload_cols.values()))
                inserted += 1
                if len(preview) < 5:
                    preview.append({"action": "inserted", "name": full_name})

        except Exception as e:
            errors.append({"row": i, "field": "", "message": str(e)})
            skipped += 1

    conn.commit()
    return {
        "matched": matched, "inserted": inserted, "skipped": skipped,
        "errors": errors, "merge_candidates": 0, "preview": preview,
    }


def process_upload(file_bytes: bytes, filename: str, entity_type: str) -> dict:
    rows = _parse_rows(file_bytes, filename)
    row_count = len(rows)

    conn = get_connection()
    try:
        if entity_type == "trials":
            result = _process_trials(rows, conn)
        elif entity_type == "organizations":
            result = _process_organizations(rows, conn)
        elif entity_type == "contacts":
            result = _process_contacts(rows, conn)
        else:
            result = {"matched": 0, "inserted": 0, "skipped": row_count,
                      "errors": [{"row": 0, "field": "entity_type", "message": f"Unknown: {entity_type}"}],
                      "merge_candidates": 0, "preview": []}
    finally:
        conn.close()

    result["row_count"] = row_count
    return result
